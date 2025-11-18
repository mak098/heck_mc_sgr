from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
import datetime
from .models import Section, AcademicYear, DocumentFolde
from teachers.models import Teacher
from affectation.models import Affectation
from student.models import Student, AttachementFile

def getAllTeacherStudentBySectionExcel(self, sections):
    academic = AcademicYear.objects.get(is_current=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Enseignants par section"

    # En-tête du fichier
    ws.merge_cells("A1:F1")
    ws["A1"] = "RDC - Ministère de l'Enseignement Supérieur et Universitaire"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])  # Ligne vide

    # En-têtes du tableau
    headers = [
        "Section",
        "Enseignant",
        "Grade",
        "Matricule Enseignant",
        "Promotion",
        "Nom Etudiant",
        "Matricule Etudiant",
        "Frais projet tutoré",
        "Frais dépôt",
        "Date retrait projet tutoré",
        "Date retrait dépôt",
    ]
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    for sec in sections:
        section = Section.objects.get(id=sec["id"])
        teachers = Teacher.objects.all().order_by("-grade__grade")
        for teacher in teachers:
            affectations = Affectation.objects.filter(
                academic_year=academic, teacher=teacher, section=section
            )
            if not affectations.exists():
                continue
            for aff in affectations:
                ws.append(
                    [
                        section.name,
                        f"{teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}".strip(),
                        str(getattr(teacher.grade, "grade", "")),
                        teacher.matricule or "",
                        getattr(aff.promotion, "code", ""),
                        aff.student,
                        aff.matricule,  # Si tu veux le nombre d'étudiants, adapte ici selon ta logique
                        aff.management_fees,  # Si tu veux le nombre d'étudiants, adapte ici selon ta logique
                        aff.deposit_fees,  # Si tu veux le nombre d'étudiants, adapte ici selon ta logique
                        aff.date_teacher_amount_collected.strftime("%Y-%m-%d") if aff.date_teacher_amount_collected else "",
                        aff.date_teacher_deposit_amount_collected.strftime("%Y-%m-%d") if aff.date_teacher_deposit_amount_collected else ""
                    ]
                )

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"affectation_{section.name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def downloadStudentsByAcademicYearExcel(self, academic_years):
    """
    Télécharge tous les étudiants appartenant aux années académiques sélectionnées
    avec les DocumentFolde comme colonnes (1 si l'étudiant a le document, 0 sinon)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Etudiants"

    # Récupérer tous les DocumentFolde pour créer les colonnes dynamiquement
    all_documents = DocumentFolde.objects.all().order_by("name")

    # Calculer le nombre total de colonnes pour le merge du header
    total_columns = 10 + all_documents.count()  # 10 colonnes de base + documents

    # En-tête du fichier
    merge_range = f"A1:{chr(64 + total_columns)}1"
    ws.merge_cells(merge_range)
    ws["A1"] = "HAUTE ECOLE DE COMMERCE DE KINSHASA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])  # Ligne vide

    # En-têtes du tableau - colonnes de base
    headers = [
        "Matricule",
        "Noms",
        "Lieu de naissance",
        "Sexe",
        "Nationalité",
        "Téléphone",
        "Email",
        "Orientation",
        "Promotion",
        "Année Académique",
    ]

    # Ajouter les DocumentFolde comme colonnes
    for doc in all_documents:
        headers.append(doc.name)

    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Parcourir toutes les années académiques sélectionnées
    for ay in academic_years:
        academic_year = AcademicYear.objects.get(year=ay["year"])

        # Récupérer tous les étudiants pour cette année académique
        students = (
            Student.objects.filter(academic_year=academic_year)
            .order_by("matricule")
            .prefetch_related("documents")
        )

        for student in students:
            # Récupérer les IDs des documents que l'étudiant possède
            student_document_ids = set(student.documents.values_list("id", flat=True))

            # Construire la ligne avec les données de base
            row = [
                student.matricule or "",
                student.names or "",
                student.date_and_place_of_birth or "",
                student.gender or "",
                student.nationality or "",
                student.phone or "",
                student.email or "",
                str(student.orientation) if student.orientation else "",
                str(student.promotion) if student.promotion else "",
                str(student.academic_year) if student.academic_year else "",
            ]

            # Ajouter 1 ou 0 pour chaque DocumentFolde
            for doc in all_documents:
                if doc.id in student_document_ids:
                    row.append(1)
                else:
                    row.append(0)

            ws.append(row)

    # Ajuster la largeur des colonnes de base
    base_column_widths = [15, 30, 25, 10, 15, 15, 25, 30, 15, 15]
    for i, width in enumerate(base_column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Ajuster la largeur des colonnes de documents (plus étroites car juste 0 ou 1)
    for i in range(len(all_documents)):
        col_letter = chr(64 + len(base_column_widths) + 1 + i)
        ws.column_dimensions[col_letter].width = 12

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Nom du fichier avec les années académiques
    years_str = "_".join([ay["year"] for ay in academic_years])
    filename = f"etudiants_{years_str}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
