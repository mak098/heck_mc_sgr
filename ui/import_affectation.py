from affectation.models import Affectation
from rest_framework.response import Response
from rest_framework import viewsets, status
import openpyxl
from parameter.models import Section, Promotion, AcademicYear
from teachers.models import Teacher
from django.http import JsonResponse


def import_excel_file(request):

    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse(
            {"error": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST
        )

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    # Handle None values in header cells
    headers = [
        (cell.value.strip().lower() if cell.value is not None else "")
        for cell in sheet[1]
    ]
    success_count = 0
    errors = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        # try:
        print(
            f">>>>>>>>>>>>>>>>>>>>>>>>teacher {row_data["teacher"]} student{row_data.get("names", "-")}",
        )
        teacher = Teacher.objects.get(matricule=row_data["teacher"])
        section = (
            Section.objects.get(id=row_data["section"])
            if row_data.get("section")
            else None
        )
        promotion = (
            Promotion.objects.get(id=row_data["promotion"])
            if row_data.get("promotion")
            else None
        )
        academic_year = AcademicYear.objects.get(year=row_data["academic_year"])
        matricule = row_data.get("matricule", "-")
        if not Affectation.objects.filter(
            teacher=teacher, matricule=matricule, academic_year=academic_year
        ).exists():
            if not Affectation.objects.filter(
                matricule=matricule, academic_year=academic_year
            ).exists():
                Affectation.objects.create(
                    teacher=teacher,
                    section=section,
                    promotion=promotion,
                    student=row_data.get("names", "-"),
                    matricule=matricule,
                    academic_year=academic_year,
                    affected_by=request.user,
                )
                success_count += 1

        # except Exception as e:
        #     errors.append({"row": row_data, "error": str(e)})

    return JsonResponse(
        {
            "message": f"{success_count} affectations importées avec succès.",
            "errors": errors,
        },
        status=status.HTTP_200_OK,
    )


def import_payement(request):

    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse(
            {"error": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST
        )

    # try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    headers = [
        (cell.value.strip().lower() if cell.value is not None else "")
        for cell in sheet[1]
    ]
    success_count = 0
    errors = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        academic_year = AcademicYear.objects.get(year=row_data["academic_year"])
        matricule = row_data.get("matricule", "-")
        if Affectation.objects.filter(
            matricule=matricule, academic_year=academic_year
        ).exists():
            affectation = Affectation.objects.filter(
                matricule=matricule, academic_year=academic_year
            ).first()

            affectation.management_fees = row_data.get("management_fees", 0.00)
            affectation.save()
            success_count += 1

    return JsonResponse(
        {
            "message": f"{success_count} affectations importées avec succès.",
            "errors": errors,
        },
        status=status.HTTP_200_OK,
    )


def import_payement_deposit_fees(request):

    import openpyxl
    from django.db import transaction

    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse(
            {"error": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST
        )

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    headers = [
        (cell.value.strip().lower() if cell.value is not None else "")
        for cell in sheet[1]
    ]
    success_count = 0
    errors = []

    # Collecte toutes les données à mettre à jour
    update_data = []
    academic_year_cache = {}

    # Pré-collecte tous les matricules et années du fichier
    matricule_annee = set()
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        rows.append(row_data)
        matricule_annee.add(
            (row_data.get("matricule", "-"), row_data.get("academic_year"))
        )

    # Précharge tous les AcademicYear nécessaires
    years = {annee for _, annee in matricule_annee}
    for year in years:
        try:
            academic_year_cache[year] = AcademicYear.objects.get(year=year)
        except AcademicYear.DoesNotExist:
            academic_year_cache[year] = None

    # Précharge toutes les affectations nécessaires
    affectations = Affectation.objects.filter(
        matricule__in=[m for m, _ in matricule_annee], academic_year__year__in=years
    )
    affectation_map = {(a.matricule, a.academic_year.year): a for a in affectations}

    # Prépare les updates
    for row_data in rows:
        matricule = row_data.get("matricule", "-")
        year = row_data.get("academic_year")
        deposit_fees = row_data.get("deposit_fees", 0.00)
        academic_year = academic_year_cache.get(year)
        if not academic_year:
            errors.append({"row": row_data, "error": "Année académique introuvable"})
            continue
        affectation = affectation_map.get((matricule, year))
        if affectation:
            affectation.deposit_fees = deposit_fees
            update_data.append(affectation)
            success_count += 1
        else:
            errors.append({"row": row_data, "error": "Affectation introuvable"})

    # Mise à jour en bulk
    if update_data:
        with transaction.atomic():
            Affectation.objects.bulk_update(update_data, ["deposit_fees"])

    return JsonResponse(
        {
            "message": f"{success_count} affectations importées avec succès.",
            "errors": errors,
        },
        status=status.HTTP_200_OK,
    )


def import_update_tuteur(request):

    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse(
            {"error": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST
        )

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    # headers = [
    #     (cell.value.strip().lower() if cell.value is not None else "") for cell in sheet[1]
    # ]
    headers = [
        (cell.value.strip().lower() if cell.value is not None else "")
        for cell in sheet[1]
    ]
    success_count = 0
    errors = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        row_data = dict(zip(headers, row))
        # print(row_data)  # Pour voir le contenu de chaque ligne
        academic_year = AcademicYear.objects.get(is_current=True)
        matricule = row_data.get("matricule etudiant", "-")
        old_teacher_matricule = row_data.get("matricule enseignant", "-")
        old_teacher = Teacher.objects.filter(matricule=old_teacher_matricule).first()
        new_teacher_matricule = row_data.get("nouveau tuteur", "-")
        new_teacher = Teacher.objects.filter(matricule=new_teacher_matricule).first()
        print(
            f">>>>{matricule}>>>> {old_teacher_matricule} >>>>{new_teacher_matricule}"
        )
        if Affectation.objects.filter(
            matricule=matricule, teacher=old_teacher, academic_year=academic_year
        ).exists():
            affectation = Affectation.objects.filter(
                matricule=matricule, teacher=old_teacher, academic_year=academic_year
            ).first()
            print(
                f">>>>>>>>>>>>>>>>>>>>>>>>teacher {affectation.matricule}",
            )
            affectation.teacher = new_teacher
            affectation.save()
            success_count += 1

    return JsonResponse(
        {
            "message": f"{success_count} affectations modifier avec succès.",
            "errors": errors,
        },
        status=status.HTTP_200_OK,
    )
