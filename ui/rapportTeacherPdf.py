from django.http import HttpResponse
from fpdf import FPDF
from io import BytesIO
import os
from django.db.models import Sum, F, FloatField
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models import Prefetch
from rest_framework import viewsets
from datetime import datetime
from parameter.models import AcademicYear, Filiere, Promotion, Firm
from student.models import Student
from teachers.models import Teacher
from affectation.models import Affectation
from django.db.models import Count
import asyncio
from parameter.models import Section
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class CustomPDF(FPDF):
    def footer(self):
        # Position à 1.5 cm du bas
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        current_time = datetime.now()
        self.cell(0, 10, f"{current_time}", 0, 0, "R")
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        # Numéro de page centré

        self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")


class ExportPdf(viewsets.ModelViewSet):
    def getAllTeacherPayementDepositSyntheseBySection(self, sections):
        from io import BytesIO
        from django.http import HttpResponse
        import os

        academic = AcademicYear.objects.get(is_current=True)
        firm = Firm.objects.all().first()
        if not firm:
            return HttpResponse("Aucune entreprise trouvée", status=404)

        for sec in sections:
            section = Section.objects.get(id=sec["id"])

            pdf = CustomPDF(orientation="P")
            pdf.add_page()

            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
            pdf.set_font("Arial", "B", 10)
            pdf.cell(
                0,
                2,
                "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
                0,
                1,
                "C",
            )

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 10, firm.name, 0, 1, "C")

            logo_path = "media/" + str(firm.logo)
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
            else:
                pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
            pdf.ln(25)
            pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

            pdf.ln(1)
            rect_width = pdf.w / 3
            rect_height = 4
            y_position = pdf.get_y()
            pdf.set_fill_color(255, 0, 0)
            pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
            pdf.set_fill_color(255, 255, 0)
            pdf.rect(
                x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD"
            )
            pdf.set_fill_color(0, 0, 255)
            pdf.rect(
                x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
            )
            pdf.ln(5)

            # Titre du rapport synthèse
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, f"{section.name}", 0, 1, "C")
            pdf.ln(1)
            pdf.cell(
                0, 10, "Répartition des frais de dépôt du projet tutoré 1", 0, 1, "C"
            )
            pdf.ln(5)
            pdf.set_fill_color(0, 0, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(10, 8, "N°", 1, 0, "C", fill=True)
            pdf.cell(50, 8, "Enseignant", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Somme Dépôt", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Perçu", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Solde", 1, 0, "C", fill=True)
            pdf.cell(25, 8, "ETS(30%)", 1, 0, "C", fill=True)
            pdf.cell(25, 8, "SGR(20%)", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "SECTION(50%)", 1, 1, "C", fill=True)

            teachers = (
                Teacher.objects.filter(
                    affectations_teacher_set__section=section,
                    affectations_teacher_set__academic_year=academic,
                )
                .distinct()
                .order_by("-grade__grade")
            )
            total_due = 0.0
            total_paid = 0.0
            total_ets = 0.0
            total_sgr = 0.0
            total_section = 0.0
            display_idx = 1

            for teacher in teachers:
                affectations = Affectation.objects.filter(
                    academic_year=academic, teacher=teacher, section=section
                )
                sum_due = sum(float(aff.deposit_fees or 0) for aff in affectations)
                sum_paid = sum(
                    float(aff.teacher_deposit_amount_collected or 0)
                    for aff in affectations
                )
                solde = sum_due - sum_paid
                ets = solde * 0.3
                sgr = solde * 0.2
                section_part = solde * 0.5

                if sum_due > 0 or sum_paid > 0:
                    total_due += sum_due
                    total_paid += sum_paid
                    total_ets += ets
                    total_sgr += sgr
                    total_section += section_part

                    teacher_name = f"{teacher.grade}. {teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}".strip()
                    pdf.cell(10, 8, str(display_idx), 1, 0, "C")
                    pdf.cell(50, 8, teacher_name, 1, 0, "L")
                    pdf.cell(30, 8, f"{sum_due:.2f}", 1, 0, "R")
                    pdf.cell(30, 8, f"{sum_paid:.2f}", 1, 0, "R")
                    pdf.cell(30, 8, f"{solde:.2f}", 1, 0, "R")
                    pdf.cell(25, 8, f"{ets:.2f}", 1, 0, "R")
                    pdf.cell(25, 8, f"{sgr:.2f}", 1, 0, "R")
                    pdf.cell(30, 8, f"{section_part:.2f}", 1, 1, "R")
                    display_idx += 1

            # Ligne des totaux
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(230, 230, 230)
            pdf.cell(60, 8, "TOTAL", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_due:.2f}", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_paid:.2f}", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_due-total_paid:.2f}", 1, 0, "R", fill=True)
            pdf.cell(25, 8, f"{total_ets:.2f}", 1, 0, "R", fill=True)
            pdf.cell(25, 8, f"{total_sgr:.2f}", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_section:.2f}", 1, 1, "R", fill=True)

            # Générer le PDF
            pdf_buffer = BytesIO()
            pdf.output(pdf_buffer, dest="S")
            pdf_buffer.seek(0)

            response = HttpResponse(pdf_buffer, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="synthese_depot_{section.name}_{academic.year}.pdf"'
            )
            return response

    def getTeacherStudent(self, year, teacher):
        try:
            if year == "current":
                academic = AcademicYear.objects.get(is_current=True)
            else:
                academic = AcademicYear.objects.get(year=year)
        except AcademicYear.DoesNotExist:
            return HttpResponse("Année académique non trouvée", status=404)

        firm = Firm.objects.all().first()
        if not firm:
            return HttpResponse("Aucune entreprise trouvée", status=404)

        pdf = CustomPDF(orientation="P")  # Portrait
        pdf.add_page()

        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
        pdf.set_font("Arial", "B", 10)
        pdf.cell(
            0,
            2,
            "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
            0,
            1,
            "C",
        )

        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 10, firm.name, 0, 1, "C")

        logo_path = "media/" + str(firm.logo)
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
        else:
            pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
        pdf.ln(25)
        pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

        pdf.ln(1)
        rect_width = pdf.w / 3
        rect_height = 4
        y_position = pdf.get_y()
        pdf.set_fill_color(255, 0, 0)
        pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(255, 255, 0)
        pdf.rect(x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(0, 0, 255)
        pdf.rect(
            x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
        )
        pdf.ln(10)

        _teacher = Teacher.objects.get(id=teacher)

        pdf.set_font("Arial", "", 10)
        pdf.cell(30, 4, "Grade :", 0, 0, "L")
        pdf.cell(40, 4, _teacher.grade.grade, 0, 1, "L")
        pdf.ln(3)
        pdf.cell(30, 4, "Matricule :", 0, 0, "L")
        pdf.cell(40, 4, _teacher.matricule, 0, 1, "L")
        pdf.ln(3)
        pdf.cell(30, 4, "Noms :", 0, 0, "L")
        pdf.cell(
            40,
            4,
            remove_non_ascii(
                f"{_teacher.first_name} {_teacher.last_name} {_teacher.name}"
            ),
            0,
            1,
            "L",
        )
        pdf.ln(3)

        affectations = Affectation.objects.filter(
            academic_year=academic, teacher=teacher
        )
        summary = (
            affectations.values("section__sigle", "promotion__code")
            .annotate(total=Count("id"))
            .order_by("section__sigle", "promotion__code")
        )
        total_general = affectations.count()

        pdf.set_font("Arial", "B", 9)
        pdf.cell(0, 8, "Récapitulatif par promotion", 0, 1, "L")
        pdf.set_fill_color(200, 200, 200)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(55, 8, "Section", 1, 0, "C", fill=True)
        pdf.cell(45, 8, "Promotion", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Nb Étudiants", 1, 1, "C", fill=True)

        pdf.set_fill_color(255, 255, 255)
        for item in summary:
            section_sigle = item["section__sigle"] or "N/A"
            promotion_code = item["promotion__code"] or "N/A"
            pdf.cell(55, 8, remove_non_ascii(section_sigle), 1, 0, "L")
            pdf.cell(45, 8, remove_non_ascii(promotion_code), 1, 0, "L")
            pdf.cell(30, 8, str(item["total"]), 1, 1, "L")

        pdf.set_fill_color(220, 220, 220)
        pdf.cell(100, 8, "TOTAL GENERAL", 1, 0, "R", fill=True)
        pdf.cell(30, 8, str(total_general), 1, 1, "R", fill=True)
        pdf.ln(5)

        pdf.set_font("Arial", "B", 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, f"Etudiants", 0, 1)

        pdf.set_fill_color(0, 0, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(15, 8, "Num", 1, 0, "C", fill=True)
        pdf.cell(60, 8, "Noms", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Section", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Promotion", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Payement", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Percu", 1, 1, "C", fill=True)

        i = 0
        total_management_fees = 0.0
        total_teacher_collected = 0.0
        pdf.set_text_color(0, 0, 0)

        for aff in affectations:
            i += 1
            m_fees = float(aff.management_fees) if aff.management_fees else 0.0
            t_collected = (
                float(aff.teacher_amount_collected)
                if aff.teacher_amount_collected
                else 0.0
            )
            total_management_fees += m_fees
            total_teacher_collected += t_collected

            pdf.cell(15, 8, str(i), 1, 0, "L")
            pdf.cell(60, 8, remove_non_ascii(aff.student), 1, 0, "L")
            pdf.cell(30, 8, remove_non_ascii(aff.section.sigle), 1, 0, "L")
            pdf.cell(
                30,
                8,
                remove_non_ascii(aff.promotion.code if aff.promotion else "N/A"),
                1,
                0,
                "L",
            )
            pdf.cell(30, 8, f"{m_fees:.2f}", 1, 0, "R")
            pdf.cell(30, 8, f"{t_collected:.2f}", 1, 1, "R")

        # Ligne des totaux
        pdf.set_font("Arial", "B", 9)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(135, 8, "TOTAL", 1, 0, "R", fill=True)
        pdf.cell(30, 8, f"{total_management_fees:.2f}", 1, 0, "R", fill=True)
        pdf.cell(30, 8, f"{total_teacher_collected:.2f}", 1, 1, "R", fill=True)

        # Ligne du disponible
        disponible = total_management_fees - total_teacher_collected
        pdf.set_fill_color(200, 255, 200)
        pdf.set_text_color(0, 100, 0)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(135, 8, "DISPONIBLE", 1, 0, "R", fill=True)
        pdf.cell(60, 8, f"{disponible:.2f}", 1, 1, "R", fill=True)

        # Générer le PDF
        pdf_buffer = BytesIO()
        pdf.output(pdf_buffer, dest="S")
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{remove_non_ascii(_teacher.first_name)}.pdf"'
        )
        return response

    # returner le rapport pour chaque prof
    def getTeacherSynthese(self, year):
        try:
            if year == "current":
                academic = AcademicYear.objects.get(is_current=True)
            else:
                academic = AcademicYear.objects.get(year=year)
        except AcademicYear.DoesNotExist:
            return HttpResponse("Année académique non trouvée", status=404)

        firm = Firm.objects.all().first()
        if not firm:
            return HttpResponse("Aucune entreprise trouvée", status=404)

        pdf = CustomPDF(orientation="P")
        pdf.add_page()

        # Titre principal
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")

        # Sous-titre
        pdf.set_text_color(0, 0, 255)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(
            0,
            2,
            "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
            0,
            1,
            "C",
        )

        # Nom de l'entreprise
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 10, firm.name, 0, 1, "C")
        pdf.cell(0, 10, firm.service, 0, 1, "C")

        # Logo sous le nom de l'entreprise avec un interligne de 10
        pdf.ln(1)
        logo_path = "media/" + str(firm.logo)
        if os.path.exists(logo_path):
            pdf.image(
                logo_path,
                x=(pdf.w / 2 - 12.5),
                y=pdf.get_y(),
                w=25,
                h=25,
                type="",
                link="",
            )
        else:
            pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")

        # Trois barres sous le logo avec un interligne de 10
        pdf.ln(28)
        rect_width = pdf.w / 3
        rect_height = 4
        y_position = pdf.get_y()

        pdf.set_fill_color(255, 0, 0)
        pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(255, 255, 0)
        pdf.rect(x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(0, 0, 255)
        pdf.rect(
            x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
        )
        pdf.ln(5)
        pdf.cell(
            0,
            10,
            "Répartition des Étudiants par Enseignant selon les Sections et Promotions",
            0,
            1,
            "C",
        )
        pdf.ln(5)
        # Récupérer tous les enseignants et calculer leur total général
        teachers = Teacher.objects.all()
        teacher_data = []
        for teacher in teachers:
            affectations = Affectation.objects.filter(
                academic_year=academic, teacher=teacher
            )
            total_general = affectations.count()

            teacher_data.append(
                {
                    "teacher": teacher,
                    "total_general": total_general,
                    "affectations": affectations,
                    "summary": (
                        affectations.values(
                            "section__sigle",
                            "promotion__code",
                        )
                        .annotate(total=Count("id"))
                        .order_by(
                            "section__sigle",
                            "promotion__code",
                        )
                    ),
                }
            )

        # Trier du plus grand au plus petit total général
        teacher_data.sort(key=lambda x: x["total_general"], reverse=True)

        for data in teacher_data:
            teacher = data["teacher"]
            total_general = data["total_general"]
            summary = data["summary"]

            # En-tête du tableau récapitulatif
            pdf.set_font("Arial", "B", 9)
            pdf.cell(
                0,
                8,
                f"{teacher.matricule or ''} {getattr(teacher, 'grade', '') or ''}. {teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}",
                0,
                1,
                "L",
            )

            pdf.set_fill_color(200, 200, 200)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(55, 8, "Section", 1, 0, "C", fill=True)
            pdf.cell(45, 8, "Promotion", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Nb Étudiants", 1, 1, "C", fill=True)

            pdf.set_fill_color(255, 255, 255)
            for item in summary:
                pdf.cell(55, 8, str(item.get("section__sigle") or ""), 1, 0, "L")
                pdf.cell(45, 8, str(item.get("promotion__code") or ""), 1, 0, "L")
                pdf.cell(30, 8, str(item.get("total") or ""), 1, 1, "R")

            pdf.set_fill_color(220, 220, 220)
            pdf.cell(100, 8, "TOTAL GENERAL", 1, 0, "R", fill=True)
            pdf.cell(30, 8, str(total_general), 1, 1, "R", fill=True)
            pdf.ln(5)

        pdf.set_font("Arial", "B", 10)
        pdf.cell(
            0,
            10,
            f"Secrétaire Général",
            0,
            1,
            "L",
        )
        pdf.ln(5)
        # Sauvegarder le PDF dans un fichier temporaire
        pdf_buffer = BytesIO()
        pdf.output(pdf_buffer, dest="S")
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{academic.year}.pdf"'
        return response

    def getAllTeacherStudent(self, year):
        try:
            if year == "current":
                academic = AcademicYear.objects.get(is_current=True)
            else:
                academic = AcademicYear.objects.get(year=year)
        except AcademicYear.DoesNotExist:
            return HttpResponse("Année académique non trouvée", status=404)

        firm = Firm.objects.all().first()
        if not firm:
            return HttpResponse("Aucune entreprise trouvée", status=404)

        pdf = CustomPDF(orientation="P")
        pdf.add_page()

        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
        pdf.set_font("Arial", "B", 10)
        pdf.cell(
            0,
            2,
            "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
            0,
            1,
            "C",
        )

        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 10, firm.name, 0, 1, "C")

        logo_path = "media/" + str(firm.logo)
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
        else:
            pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
        pdf.ln(25)
        pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

        pdf.ln(1)
        rect_width = pdf.w / 3
        rect_height = 4
        y_position = pdf.get_y()
        pdf.set_fill_color(255, 0, 0)
        pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(255, 255, 0)
        pdf.rect(x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(0, 0, 255)
        pdf.rect(
            x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
        )
        pdf.ln(10)

        teachers = Teacher.objects.all()
        for _teacher in teachers:
            pdf.set_font("Arial", "B", 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(
                0,
                10,
                f"Enseignant : {(_teacher.first_name or '')} {(_teacher.last_name or '')} {(_teacher.name or '')}",
                0,
                1,
                "L",
            )
            pdf.set_font("Arial", "", 10)
            pdf.cell(30, 4, "Grade :", 0, 0, "L")
            pdf.cell(40, 4, str(getattr(_teacher.grade, "grade", "")), 0, 1, "L")
            pdf.ln(3)
            pdf.cell(30, 4, "Matricule :", 0, 0, "L")
            pdf.cell(40, 4, str(_teacher.matricule or ""), 0, 1, "L")
            pdf.ln(3)

            affectations = Affectation.objects.filter(
                academic_year=academic, teacher=_teacher
            )
            if not affectations.exists():
                pdf.cell(0, 8, "Aucune affectation pour cet enseignant.", 0, 1, "L")
                pdf.ln(5)
                continue

            pdf.set_font("Arial", "B", 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 10, f"Etudiants", 0, 1)

            pdf.set_fill_color(0, 0, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", "B", 9)
            pdf.cell(15, 8, "Num", 1, 0, "C", fill=True)
            pdf.cell(60, 8, "Noms", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Section", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Promotion", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Payement", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Percu", 1, 1, "C", fill=True)

            i = 0
            total_management_fees = 0.0
            total_teacher_collected = 0.0
            pdf.set_text_color(0, 0, 0)

            for aff in affectations:
                i += 1
                m_fees = float(aff.management_fees) if aff.management_fees else 0.0
                t_collected = (
                    float(aff.teacher_amount_collected)
                    if aff.teacher_amount_collected
                    else 0.0
                )
                total_management_fees += m_fees
                total_teacher_collected += t_collected

                pdf.cell(15, 8, str(i), 1, 0, "L")
                pdf.cell(60, 8, aff.student, 1, 0, "L")
                pdf.cell(30, 8, aff.section.sigle, 1, 0, "L")
                pdf.cell(
                    30, 8, aff.promotion.code if aff.promotion else "N/A", 1, 0, "L"
                )
                pdf.cell(30, 8, f"{m_fees:.2f}", 1, 0, "R")
                pdf.cell(30, 8, f"{t_collected:.2f}", 1, 1, "R")
            # Ligne des totaux
            pdf.set_font("Arial", "B", 9)
            pdf.set_fill_color(230, 230, 230)
            pdf.cell(135, 8, "TOTAL", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_management_fees:.2f}", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_teacher_collected:.2f}", 1, 1, "R", fill=True)

            # Ligne du disponible
            disponible = total_management_fees - total_teacher_collected
            pdf.set_fill_color(200, 255, 200)
            pdf.set_text_color(0, 100, 0)
            pdf.set_font("Arial", "B", 9)
            pdf.cell(135, 8, "DISPONIBLE", 1, 0, "R", fill=True)
            pdf.cell(60, 8, f"{disponible:.2f}", 1, 1, "R", fill=True)
            pdf.ln(10)

        # Générer le PDF
        pdf_buffer = BytesIO()
        pdf.output(pdf_buffer, dest="S")
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="all_teachers_{academic.year}.pdf"'
        )
        return response

    def getAllTeacherStudentBySection(self, sections):
        year = "current"
        for sec in sections:
            section_id = Section.objects.get(id=sec["id"])
            try:
                if year == "current":
                    academic = AcademicYear.objects.get(is_current=True)
                else:
                    academic = AcademicYear.objects.get(year=year)
            except AcademicYear.DoesNotExist:
                return HttpResponse("Année académique non trouvée", status=404)

            firm = Firm.objects.all().first()
            if not firm:
                return HttpResponse("Aucune entreprise trouvée", status=404)

            pdf = CustomPDF(orientation="P")
            pdf.add_page()

            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
            pdf.set_font("Arial", "B", 10)
            pdf.cell(
                0,
                2,
                "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
                0,
                1,
                "C",
            )

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 10, firm.name, 0, 1, "C")

            logo_path = "media/" + str(firm.logo)
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
            else:
                pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
            pdf.ln(25)
            pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

            pdf.ln(1)
            rect_width = pdf.w / 3
            rect_height = 4
            y_position = pdf.get_y()
            pdf.set_fill_color(255, 0, 0)
            pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
            pdf.set_fill_color(255, 255, 0)
            pdf.rect(
                x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD"
            )
            pdf.set_fill_color(0, 0, 255)
            pdf.rect(
                x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
            )
            pdf.ln(10)

            pdf.cell(0, 10, f"{section_id.name}", 0, 1, "C")

            # Filtrer les enseignants qui ont des affectations dans cette section
            teachers = Teacher.objects.filter(
                affectations_teacher_set__academic_year=academic,
                affectations_teacher_set__section_id=section_id,
            ).distinct()

            for _teacher in teachers:
                pdf.set_font("Arial", "B", 12)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(
                    0,
                    10,
                    f"Enseignant : {(_teacher.first_name or '')} {(_teacher.last_name or '')} {(_teacher.name or '')}",
                    0,
                    1,
                    "L",
                )
                pdf.set_font("Arial", "", 10)
                pdf.cell(30, 4, "Grade :", 0, 0, "L")
                pdf.cell(40, 4, str(getattr(_teacher.grade, "grade", "")), 0, 1, "L")
                pdf.ln(3)
                pdf.cell(30, 4, "Matricule :", 0, 0, "L")
                pdf.cell(40, 4, str(_teacher.matricule or ""), 0, 1, "L")
                pdf.ln(3)

                affectations = Affectation.objects.filter(
                    academic_year=academic, teacher=_teacher, section_id=section_id
                )
                if not affectations.exists():
                    pdf.cell(
                        0,
                        8,
                        "Aucune affectation pour cet enseignant dans cette section.",
                        0,
                        1,
                        "L",
                    )
                    pdf.ln(5)
                    continue

                pdf.set_font("Arial", "B", 12)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(0, 10, f"Etudiants", 0, 1)

                pdf.set_fill_color(0, 0, 0)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font("Arial", "B", 9)
                pdf.cell(15, 8, "Num", 1, 0, "C", fill=True)
                pdf.cell(60, 8, "Noms", 1, 0, "C", fill=True)
                pdf.cell(30, 8, "Section", 1, 0, "C", fill=True)
                pdf.cell(30, 8, "Promotion", 1, 0, "C", fill=True)
                pdf.cell(30, 8, "Payement", 1, 0, "C", fill=True)
                pdf.cell(30, 8, "Percu", 1, 1, "C", fill=True)

                i = 0
                total_management_fees = 0.0
                total_teacher_collected = 0.0
                pdf.set_text_color(0, 0, 0)

                for aff in affectations:
                    i += 1
                    m_fees = float(aff.management_fees) if aff.management_fees else 0.0
                    t_collected = (
                        float(aff.teacher_amount_collected)
                        if aff.teacher_amount_collected
                        else 0.0
                    )
                    total_management_fees += m_fees
                    total_teacher_collected += t_collected

                    pdf.cell(15, 8, str(i), 1, 0, "L")
                    pdf.cell(60, 8, remove_non_ascii(aff.student), 1, 0, "L")
                    pdf.cell(
                        30,
                        8,
                        remove_non_ascii(getattr(aff.section, "sigle", "")),
                        1,
                        0,
                        "L",
                    )
                    pdf.cell(
                        30,
                        8,
                        remove_non_ascii(getattr(aff.promotion, "code", "N/A")),
                        1,
                        0,
                        "L",
                    )
                    pdf.cell(30, 8, f"{m_fees:.2f}", 1, 0, "R")
                    pdf.cell(30, 8, f"{t_collected:.2f}", 1, 1, "R")
                # Ligne des totaux
                pdf.set_font("Arial", "B", 9)
                pdf.set_fill_color(230, 230, 230)
                pdf.cell(135, 8, "TOTAL", 1, 0, "R", fill=True)
                pdf.cell(30, 8, f"{total_management_fees:.2f}", 1, 0, "R", fill=True)
                pdf.cell(30, 8, f"{total_teacher_collected:.2f}", 1, 1, "R", fill=True)

                # Ligne du disponible
                disponible = total_management_fees - total_teacher_collected
                pdf.set_fill_color(200, 255, 200)
                pdf.set_text_color(0, 100, 0)
                pdf.set_font("Arial", "B", 9)
                pdf.cell(135, 8, "DISPONIBLE", 1, 0, "R", fill=True)
                pdf.cell(60, 8, f"{disponible:.2f}", 1, 1, "R", fill=True)
                pdf.ln(10)

            # Générer le PDF
        pdf_buffer = BytesIO()
        pdf.output(pdf_buffer, dest="S")
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="teachers_section_{section_id}_{academic.year}.pdf"'
        )
        return response

    def getAllTeacherPayementSynthese(self, year):
        try:
            if year == "current":
                academic = AcademicYear.objects.get(is_current=True)
            else:
                academic = AcademicYear.objects.get(year=year)
        except AcademicYear.DoesNotExist:
            return HttpResponse("Année académique non trouvée", status=404)

        firm = Firm.objects.all().first()
        if not firm:
            return HttpResponse("Aucune entreprise trouvée", status=404)

        pdf = CustomPDF(orientation="P")
        pdf.add_page()

        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
        pdf.set_font("Arial", "B", 10)
        pdf.cell(
            0,
            2,
            "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
            0,
            1,
            "C",
        )

        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 10, firm.name, 0, 1, "C")

        logo_path = "media/" + str(firm.logo)
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
        else:
            pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
        pdf.ln(25)
        pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

        pdf.ln(1)
        rect_width = pdf.w / 3
        rect_height = 4
        y_position = pdf.get_y()
        pdf.set_fill_color(255, 0, 0)
        pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(255, 255, 0)
        pdf.rect(x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD")
        pdf.set_fill_color(0, 0, 255)
        pdf.rect(
            x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
        )
        pdf.ln(10)

        # Titre du rapport synthèse
        pdf.set_font("Arial", "B", 12)
        pdf.cell(
            0, 10, "État de Paiement des Enseignants du Programme Tutoré", 0, 1, "C"
        )
        pdf.ln(5)

        # En-tête du tableau
        pdf.set_fill_color(0, 0, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(10, 8, "N°", 1, 0, "C", fill=True)
        pdf.cell(80, 8, "Enseignant", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Somme Paiement", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Perçu", 1, 0, "C", fill=True)
        pdf.cell(30, 8, "Solde", 1, 1, "C", fill=True)

        teachers = Teacher.objects.all().order_by("-grade")
        pdf.set_font("Arial", "", 10)
        pdf.set_text_color(0, 0, 0)
        total_due = 0.0
        total_paid = 0.0

        for idx, teacher in enumerate(teachers, start=1):
            affectations = Affectation.objects.filter(
                academic_year=academic, teacher=teacher
            )
            sum_due = sum(float(aff.management_fees or 0) for aff in affectations)
            sum_paid = sum(
                float(aff.teacher_amount_collected or 0) for aff in affectations
            )
            total_due += sum_due
            total_paid += sum_paid

            teacher_name = f"{teacher.grade}. {teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}".strip()
            pdf.cell(10, 8, str(idx), 1, 0, "C")
            pdf.cell(80, 8, teacher_name, 1, 0, "L")
            pdf.cell(30, 8, f"{sum_due:.2f}", 1, 0, "R")
            pdf.cell(30, 8, f"{sum_paid:.2f}", 1, 0, "R")
            dispo = sum_due - sum_paid
            pdf.cell(30, 8, f"{dispo:.2f}", 1, 1, "R")

        # Ligne des totaux
        pdf.set_font("Arial", "B", 10)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(90, 8, "TOTAL", 1, 0, "R", fill=True)
        pdf.cell(30, 8, f"{total_due:.2f}", 1, 0, "R", fill=True)
        pdf.cell(30, 8, f"{total_paid:.2f}", 1, 0, "R", fill=True)
        disponible = total_due - total_paid
        pdf.cell(30, 8, f"{disponible:.2f}", 1, 1, "R", fill=True)
        # Générer le PDF
        pdf_buffer = BytesIO()
        pdf.output(pdf_buffer, dest="S")
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="synthese_paiement_enseignants_{academic.year}.pdf"'
        )
        return response

    def getAllTeacherPayementSyntheseBySection(self, sections):
        from io import BytesIO
        from django.http import HttpResponse
        import os

        academic = AcademicYear.objects.get(is_current=True)

        for sec in sections:
            section = Section.objects.get(id=sec["id"])
            firm = Firm.objects.all().first()
            if not firm:
                return HttpResponse("Aucune entreprise trouvée", status=404)

            pdf = CustomPDF(orientation="P")
            pdf.add_page()

            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
            pdf.set_font("Arial", "B", 10)
            pdf.cell(
                0,
                2,
                "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
                0,
                1,
                "C",
            )

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 10, firm.name, 0, 1, "C")

            logo_path = "media/" + str(firm.logo)
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
            else:
                pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
            pdf.ln(25)
            pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

            pdf.ln(1)
            rect_width = pdf.w / 3
            rect_height = 4
            y_position = pdf.get_y()
            pdf.set_fill_color(255, 0, 0)
            pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
            pdf.set_fill_color(255, 255, 0)
            pdf.rect(
                x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD"
            )
            pdf.set_fill_color(0, 0, 255)
            pdf.rect(
                x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
            )
            pdf.ln(5)

            # Titre du rapport synthèse
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, f"{section.name}", 0, 1, "C")
            pdf.ln(1)
            pdf.cell(
                0, 10, "État de Paiement des Enseignants du Programme Tutoré", 0, 1, "C"
            )
            pdf.ln(5)

            # En-tête du tableau
            pdf.set_fill_color(0, 0, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(10, 8, "N°", 1, 0, "C", fill=True)
            pdf.cell(80, 8, "Enseignant", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Somme Paiement", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Perçu", 1, 0, "C", fill=True)
            pdf.cell(30, 8, "Solde", 1, 1, "C", fill=True)

            teachers = Teacher.objects.all().order_by("-grade__grade")
            pdf.set_font("Arial", "", 10)
            pdf.set_text_color(0, 0, 0)
            total_due = 0.0
            total_paid = 0.0
            display_idx = 1

            for teacher in teachers:
                affectations = Affectation.objects.filter(
                    academic_year=academic, teacher=teacher, section=section
                )
                sum_due = sum(float(aff.management_fees or 0) for aff in affectations)
                sum_paid = sum(
                    float(aff.teacher_amount_collected or 0) for aff in affectations
                )

                if sum_due > 0:
                    total_due += sum_due
                    total_paid += sum_paid

                    teacher_name = f"{teacher.grade}. {teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}".strip()
                    pdf.cell(10, 8, str(display_idx), 1, 0, "C")
                    pdf.cell(80, 8, teacher_name, 1, 0, "L")
                    pdf.cell(30, 8, f"{sum_due:.2f}", 1, 0, "R")
                    pdf.cell(30, 8, f"{sum_paid:.2f}", 1, 0, "R")
                    dispo = sum_due - sum_paid
                    pdf.cell(30, 8, f"{dispo:.2f}", 1, 1, "R")
                    display_idx += 1

            # Ligne des totaux
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(230, 230, 230)
            pdf.cell(90, 8, "TOTAL", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_due:.2f}", 1, 0, "R", fill=True)
            pdf.cell(30, 8, f"{total_paid:.2f}", 1, 0, "R", fill=True)
            disponible = total_due - total_paid
            pdf.cell(30, 8, f"{disponible:.2f}", 1, 1, "R", fill=True)

            # Générer le PDF
            pdf_buffer = BytesIO()
            pdf.output(pdf_buffer, dest="S")
            pdf_buffer.seek(0)

            response = HttpResponse(pdf_buffer, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="synthese_paiement_{section.name}_{academic.year}.pdf"'
            )
            return response

    # def getAllTeacherPayementSyntheseBySection(self, sections):
    #     from io import BytesIO
    #     from django.http import HttpResponse
    #     import os

    #     academic = AcademicYear.objects.get(is_current=True)

    #     for sec in sections:
    #         section = Section.objects.get(id=sec["id"])
    #         firm = Firm.objects.all().first()
    #         if not firm:
    #             return HttpResponse("Aucune entreprise trouvée", status=404)

    #         pdf = CustomPDF(orientation="P")
    #         pdf.add_page()

    #         pdf.set_font("Arial", "B", 14)
    #         pdf.cell(0, 10, "Republique democratique du Congo".upper(), 0, 1, "C")
    #         pdf.set_font("Arial", "B", 10)
    #         pdf.cell(
    #             0,
    #             2,
    #             "Ministère de l'Enseignement Supérieur et Universitaire".upper(),
    #             0,
    #             1,
    #             "C",
    #         )

    #         pdf.set_text_color(0, 0, 0)
    #         pdf.set_font("Arial", "B", 10)
    #         pdf.cell(0, 10, firm.name, 0, 1, "C")

    #         logo_path = "media/" + str(firm.logo)
    #         if os.path.exists(logo_path):
    #             pdf.image(logo_path, x=(pdf.w / 2 - 12.5), y=pdf.get_y(), w=25, h=25)
    #         else:
    #             pdf.cell(0, 10, "Logo non trouvé", 0, 1, "C")
    #         pdf.ln(25)
    #         pdf.cell(0, 10, firm.service.upper(), 0, 1, "C")

    #         pdf.ln(1)
    #         rect_width = pdf.w / 3
    #         rect_height = 4
    #         y_position = pdf.get_y()
    #         pdf.set_fill_color(255, 0, 0)
    #         pdf.rect(x=0, y=y_position, w=rect_width, h=rect_height, style="FD")
    #         pdf.set_fill_color(255, 255, 0)
    #         pdf.rect(
    #             x=rect_width, y=y_position, w=rect_width, h=rect_height, style="FD"
    #         )
    #         pdf.set_fill_color(0, 0, 255)
    #         pdf.rect(
    #             x=rect_width * 2, y=y_position, w=rect_width, h=rect_height, style="FD"
    #         )
    #         pdf.ln(5)

    #         # Titre du rapport synthèse
    #         pdf.set_font("Arial", "B", 12)
    #         pdf.cell(0, 10, f"{section.name}", 0, 1, "C")
    #         pdf.ln(1)
    #         pdf.cell(
    #             0, 10, "État de Paiement des Enseignants du Programme Tutoré", 0, 1, "C"
    #         )
    #         pdf.ln(5)

    #         # En-tête du tableau
    #         pdf.set_fill_color(0, 0, 0)
    #         pdf.set_text_color(255, 255, 255)
    #         pdf.set_font("Arial", "B", 10)
    #         pdf.cell(10, 8, "N°", 1, 0, "C", fill=True)
    #         pdf.cell(80, 8, "Enseignant", 1, 0, "C", fill=True)
    #         pdf.cell(30, 8, "Somme Paiement", 1, 0, "C", fill=True)
    #         pdf.cell(30, 8, "Perçu", 1, 0, "C", fill=True)
    #         pdf.cell(30, 8, "Solde", 1, 1, "C", fill=True)

    #         teachers = Teacher.objects.all().order_by("-grade__grade")
    #         pdf.set_font("Arial", "", 10)
    #         pdf.set_text_color(0, 0, 0)
    #         total_due = 0.0
    #         total_paid = 0.0
    #         display_idx = 1

    #         for teacher in teachers:
    #             affectations = Affectation.objects.filter(
    #                 academic_year=academic, teacher=teacher, section=section
    #             )
    #             sum_due = sum(float(aff.management_fees or 0) for aff in affectations)
    #             sum_paid = sum(
    #                 float(aff.teacher_amount_collected or 0) for aff in affectations
    #             )

    #             if sum_due > 0:
    #                 total_due += sum_due
    #                 total_paid += sum_paid

    #                 teacher_name = f"{teacher.grade}. {teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}".strip()
    #                 pdf.cell(10, 8, str(display_idx), 1, 0, "C")
    #                 pdf.cell(80, 8, teacher_name, 1, 0, "L")
    #                 pdf.cell(30, 8, f"{sum_due:.2f}", 1, 0, "R")
    #                 pdf.cell(30, 8, f"{sum_paid:.2f}", 1, 0, "R")
    #                 dispo = sum_due - sum_paid
    #                 pdf.cell(30, 8, f"{dispo:.2f}", 1, 1, "R")
    #                 display_idx += 1

    #         # Ligne des totaux
    #         pdf.set_font("Arial", "B", 10)
    #         pdf.set_fill_color(230, 230, 230)
    #         pdf.cell(90, 8, "TOTAL", 1, 0, "R", fill=True)
    #         pdf.cell(30, 8, f"{total_due:.2f}", 1, 0, "R", fill=True)
    #         pdf.cell(30, 8, f"{total_paid:.2f}", 1, 0, "R", fill=True)
    #         disponible = total_due - total_paid
    #         pdf.cell(30, 8, f"{disponible:.2f}", 1, 1, "R", fill=True)

    #         # Générer le PDF
    #         pdf_buffer = BytesIO()
    #         pdf.output(pdf_buffer, dest="S")
    #         pdf_buffer.seek(0)

    #         response = HttpResponse(pdf_buffer, content_type="application/pdf")
    #         response["Content-Disposition"] = (
    #             f'attachment; filename="synthese_paiement_{section.name}_{academic.year}.pdf"'
    #         )
    #         return response

    def getAllTeacherPayementSyntheseBySectionExcel(self, sections):
        from django.http import HttpResponse

        academic = AcademicYear.objects.get(is_current=True)

        for sec in sections:
            section = Section.objects.get(id=sec["id"])
            firm = Firm.objects.all().first()
            if not firm:
                return HttpResponse("Aucune entreprise trouvée", status=404)

            wb = Workbook()
            ws = wb.active
            ws.title = f"Synthèse Paiement {section.name}"

            # Entête
            ws.merge_cells("A1:E1")
            ws["A1"] = "Republique democratique du Congo".upper()
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:E2")
            ws["A2"] = "Ministère de l'Enseignement Supérieur et Universitaire".upper()
            ws["A2"].font = Font(bold=True, size=10, color="0000FF")
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A3:E3")
            ws["A3"] = firm.name
            ws["A3"].font = Font(bold=True, size=10)
            ws["A3"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A4:E4")
            ws["A4"] = firm.service.upper()
            ws["A4"].font = Font(bold=True, size=10)
            ws["A4"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A5:E5")
            ws["A5"] = f"{section.name}"
            ws["A5"].font = Font(bold=True, size=12)
            ws["A5"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A6:E6")
            ws["A6"] = "État de Paiement des Enseignants du Programme Tutoré"
            ws["A6"].font = Font(bold=True, size=12)
            ws["A6"].alignment = Alignment(horizontal="center")

            # En-tête du tableau
            headers = ["N°", "Enseignant", "Somme Paiement", "Perçu", "Solde"]
            ws.append(headers)
            for col in range(1, 6):
                ws.cell(row=7, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=7, column=col).fill = PatternFill("solid", fgColor="000000")
                ws.cell(row=7, column=col).alignment = Alignment(horizontal="center")

            teachers = Teacher.objects.all().order_by("-grade__grade")
            total_due = 0.0
            total_paid = 0.0
            display_idx = 1
            row_idx = 8

            for teacher in teachers:
                affectations = Affectation.objects.filter(
                    academic_year=academic, teacher=teacher, section=section
                )
                sum_due = sum(float(aff.management_fees or 0) for aff in affectations)
                sum_paid = sum(
                    float(aff.teacher_amount_collected or 0) for aff in affectations
                )

                if sum_due > 0:
                    total_due += sum_due
                    total_paid += sum_paid

                    teacher_name = f"{teacher.grade}. {teacher.first_name or ''} {teacher.last_name or ''} {teacher.name or ''}".strip()
                    ws.append(
                        [
                            display_idx,
                            teacher_name,
                            f"{sum_due:.2f}",
                            f"{sum_paid:.2f}",
                            f"{sum_due - sum_paid:.2f}",
                        ]
                    )
                    row_idx += 1
                    display_idx += 1

            # Ligne des totaux
            ws.append(
                [
                    "TOTAL",
                    "",
                    f"{total_due:.2f}",
                    f"{total_paid:.2f}",
                    f"{total_due - total_paid:.2f}",
                ]
            )
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    "solid", fgColor="E6E6E6"
                )
                ws.cell(row=row_idx, column=col).alignment = Alignment(
                    horizontal="center" if col == 1 else "right"
                )

            # Ajuster la largeur des colonnes
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 18

            # Générer le fichier Excel
            from io import BytesIO

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                f'attachment; filename="synthese_paiement_{section.name}_{academic.year}.xlsx"'
            )
            return response


import unicodedata


def remove_non_ascii(text):
    if not text:
        return ""
    # Remplace les apostrophes typographiques par des simples
    text = text.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
    # Enlève les caractères non-ASCII
    return (
        unicodedata.normalize("NFKD", str(text))
        .encode("ascii", "ignore")
        .decode("ascii")
    )
